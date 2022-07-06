import logging
import inspect
import softest
import csv
from openpyxl import Workbook,load_workbook

class Utility(softest.TestCase):
	
	def get_items_text(self,items):
		item_list = []
		for item in items:
			item_list.append(item.text)
		return item_list

	def display_data(self,data):
		loop = 1
		for item in data:
			self.soft_assert(self.assertTrue,len(item)>5)
			if len(item)>5:
				print("{}. {} | Passed".format(loop,item))
			else:
				print("{}. {} | Failed".format(loop,item))
			loop = loop+1
		self.assert_all()

	@classmethod
	def custom_logger(self,logging_level=logging.DEBUG):
		#create logger
		#print(inspect.stack())
		logger_name = inspect.stack()[1][1]
		logger = logging.getLogger(logger_name)
		logger.setLevel(logging_level)
		#create fomatter
		formatter = logging.Formatter("%(asctime)s - %(levelname)s - %(name)s: %(message)s",datefmt="%d/%m/%Y %I:%M:%S %p")
		#handler
		fh = logging.FileHandler("./reports/logs/project_log_file.log")
		#set formatter
		fh.setFormatter(formatter)
		#add handler
		logger.addHandler(fh)
		return logger

	def get_excel_data(self,file_name,sheet_name):
		wb = load_workbook(file_name)
		sh = wb[sheet_name]
		rows = sh.max_row
		cols = sh.max_column
		data = []
		for i in range(2,rows+1):
			row = []
			for j in range(1,cols+1):
				row.append(sh.cell(row=i,column=j).value)
			data.append(row)
		return data

	def get_csv_data(self,file_name):
		csvdata = open(file_name,"r")
		reader = csv.reader(csvdata)
		next(reader)
		data = []
		for row in reader:
			data.append(row)
		return data
